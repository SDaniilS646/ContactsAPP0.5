from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries
from shutil import copyfile

import gc

def create_xlList():

  # ====================================================
  # КОПИРОВАНИЕ ШАБЛОНА
  # ====================================================

  template = "sender_template.xlsm"
  xl_output = "new_file.xlsm"
  copyfile(template, xl_output)

  # ====================================================
  # ОТКРЫТИЕ ФАЙЛА, ПОДГОТОВКА ПЕРЕМЕННЫХ И ОБЪЕКТОВ
  # ====================================================

  wb = load_workbook(xl_output, keep_vba=True)
  ws = wb["Рассылка"]

  lo = ws.tables

  table = lo["Рассылка"]
  table_adr = table.ref

  min_row, min_col, max_col, max_row = range_boundaries(table_adr)

  # ====================================================
  # СОХРАНЕНИЕ СТИЛЕЙ СТРОКИ
  # ====================================================

  row_style = {}

  for col in range(min_col, max_col + 1):
    row_style[col] = ws.cell(min_row + 1, col)._style

  # ====================================================
  # ПРОСТАВЛЕНИЕ ЗАЧЕНИЙ
  # ====================================================

  for i in range(2, 10):
    # ЗНАЧЕНИЯ
    ws.cell(i, 2).value = "ТЕСТ"

    # ПРОТЯГИВАНИЕ ТАБЛИЦЫ
    table.ref = f"A1:N{ws.max_row}"

    # ПРОСТАВЛЕНИЕ СТИЛЕЙ В СТРОКЕ
    for col in range(min_col, max_col + 1):
      ws.cell(i, col)._style = row_style[col]

  # ====================================================
  # СОХРАНЕНИЕ И ЗАКРЫТИЕ ФАЙЛА
  # ====================================================

  wb.save("new_file.xlsm")
  wb.close()
  gc.collect()
