from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries
from shutil import copyfile
from io import BytesIO

from pathlib import Path

import gc

TEMPLATE_PATH = Path(__file__).parent / 'sender_template.xlsm'

def create_sendfile(output_data=None):
  if not output_data:
    return None

  # ====================================================
  # ЧТЕНИЕ ШАБЛОНА
  # ====================================================

  with open(TEMPLATE_PATH, 'rb') as f:
    template_bytes = f.read()

  # ====================================================
  # ОТКРЫТИЕ ФАЙЛА, ПОДГОТОВКА ПЕРЕМЕННЫХ И ОБЪЕКТОВ
  # ====================================================

  buffer = BytesIO(template_bytes)
  wb = load_workbook(buffer, keep_vba=True)
  ws = wb["Рассылка"]

  table = ws.tables["Рассылка"]

  min_col, min_row, max_col, _ = range_boundaries(table.ref)

  # ====================================================
  # СОХРАНЕНИЕ СТИЛЕЙ СТРОКИ И ЗАГОЛОВКОВ СТОЛБЦОВ
  # ====================================================

  row_style = {}
  column_names = {}

  table.tableStyleInfo = None

  for col in range(min_col, max_col + 1):
    row_style[col] = ws.cell(min_row + 1, col)._style
    column_names[ws.cell(min_row, col).value] = col

  # ====================================================
  # ПРОСТАВЛЕНИЕ ЗАЧЕНИЙ
  # ====================================================

  start_row = min_row + 1

  for i, row_values in enumerate(output_data):
    current_row = start_row + i
    ws.cell(current_row, get_column(column_names, 'п/п')).value = i + 1
    ws.cell(current_row, get_column(column_names, 'Компания')).value = row_values.get('company_name', '')
    ws.cell(current_row, get_column(column_names, 'Контакты')).value = row_values.get('names', '')
    ws.cell(current_row, get_column(column_names, 'Почты')).value = row_values.get('mails', '')
    ws.cell(current_row, get_column(column_names, 'Телефоны')).value = row_values.get('phones', '')
    ws.cell(current_row, get_column(column_names, 'Примечание')).value = row_values.get('materials', '')


    for col_idx in range(min_col, max_col + 1):
      cell = ws.cell(current_row, col_idx)
      cell._style = row_style.get(col_idx, row_style[min_col])

  table.ref = (
    f"{ws.cell(min_row, min_col).coordinate}:"
    f"{ws.cell(ws.max_row, max_col).coordinate}"
  )
  # ====================================================
  # СОХРАНЕНИЕ И ЗАКРЫТИЕ ФАЙЛА
  # ====================================================

  output_buffer = BytesIO()
  wb.save(output_buffer)
  # gc.collect()
  output_buffer.seek(0)
  buffer.close()

  return output_buffer

def get_column(column_names, header_name):
  try:
    return column_names[header_name]
  except KeyError:
    return ValueError(
      f"В шаблоне не  найден столбец с заголовком '{header_name}. "
      f"Найденные заголовки: {list(column_names.keys())}"
    )